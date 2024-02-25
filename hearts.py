from openpyxl import Workbook, load_workbook

def Hearts_function(top, bottom, left, right, mode, target_value):
    print(top)
    print(bottom)
    print(left)
    print(right)
    top_points = 0
    bottom_points = 0
    left_points = 0
    right_points = 0
    rounds=1
    passing = ["Center","Left","Right"]

    
    while(True):
        current_pattern = passing[rounds % len(passing)]

        wb = load_workbook(filename='GameGenie/Hearts.xlsx')
        ws = wb["Sheet1"]
        #Showing Passing Order
        print(current_pattern)#i want this to be on page 2 in info box
        input("Press enter to continue...") #I want this to become a button you press on page 2
        #Rounds
        print(rounds)#I want this to be on page 2 in info box
        #please import the chart to the pyside 2nd page so It can be edited
        wb.save('GameGenie/Hearts.xlsx')
        temp_score = f"A{rounds}"
        top_points += int(ws[temp_score].value)
        temp_score = f"B{rounds}"
        bottom_points += int(ws[temp_score].value)
        temp_score = f"C{rounds}"
        right_points += int(ws[temp_score].value)
        temp_score = f"D{rounds}"
        left_points += int(ws[temp_score].value)

        #Showing total scores
        print(top_points) #I want This to display in a 1*4 chart in position 1
        print(bottom_points) #I want this to display in the same chart at possition 2
        print(left_points)#I want this to display in the same chart at possition 3
        print(right_points)#I want this to display in the same chart at possition 4

        if mode:
            if (top_points >=target_value and bottom_points >=target_value and left_points >=target_value and right_points >=target_value):
                break
        else:
            if (rounds > target_value):
                break
        rounds += 1
    

Hearts_function("eli", "alex", "Grilliot", "Tom",True, 10)
